import serial  # 引用pySerial模組
import socket
import time
import numpy as np
import matplotlib.pyplot as plt
import datetime
import os
import keras
from keras.preprocessing import image
from keras.models import Model
from keras import models
from keras import layers
from keras.applications.resnet50 import ResNet50
import sys
import cv2
from tensorflow.keras import layers
from tensorflow import keras
from sklearn import datasets, metrics
from sklearn.neighbors import KNeighborsClassifier
from sklearn.preprocessing import StandardScaler
from sklearn.model_selection import train_test_split
from sklearn.svm import LinearSVC
from sklearn.metrics import accuracy_score
from keras.models import Sequential, model_from_json
from keras.layers import Input, Flatten, Dense
from openpyxl import Workbook
from skimage import io
from PIL import Image
from enum import Enum
import numpy as np
import pandas as pd
import random
import joblib
import csv


COM_PORT = 'COM6'  # 指定通訊埠名稱
BAUD_RATES = 9600  # 設定傳輸速率
ser = serial.Serial(COM_PORT, BAUD_RATES)  # 初始化序列通訊埠
arr = []
IMAGE_SIZE = 640
path = 'D:/chung_yuan/vibration/bottle/'
Max = 0
HOST = '192.168.43.27'  # 設定要綁定的地址
PORT = 8001

# 按照指定圖片大小調整尺寸
def resize_image(waveImage, height=IMAGE_SIZE, width=IMAGE_SIZE):
    top, bottom, left, right = (0, 0, 0, 0)

    # 獲取圖片尺寸
    h, w, _ = waveImage.shape

    # 對於長寬不相等的圖片，找到最長的一邊
    longest_edge = max(h, w)

    # 計算短邊需要增加多上像素寬度使其與邊長等長
    if h < longest_edge:
        dh = longest_edge - h
        top = dh // 2
        bottom = dh - top
    elif w < longest_edge:
        dw = longest_edge - w
        left = dw // 2
        right = dw - left
    else:
        pass

	# RGB颜色
    WHITE = [255, 255, 255]

    # 給圖片增加邊界，使圖片長、寬等長，cv2.BORDER_CONSTANT指定邊界颜色由value指定
    constant = cv2.copyMakeBorder(waveImage, top, bottom, left, right, cv2.BORDER_CONSTANT, value=WHITE)

    # 調整圖片大小並返回
    return cv2.resize(constant, (height, width))


# 讀取訓練數據
images = []
labels = []


def read_path(path_name):
    for dir_item in os.listdir(path_name):
        # 從初始路徑開始疊加，合併成可以識別的操作路徑
        full_path = os.path.abspath(os.path.join(path_name, dir_item))

        if os.path.isdir(full_path):  # 如果是資料夾，繼續遞迴調用
            read_path(full_path)
        else:  # 文件
            if dir_item.endswith('.png'):
                waveImage = cv2.imread(full_path)
                waveImage = resize_image(waveImage, IMAGE_SIZE, IMAGE_SIZE)
				
                images.append(waveImage)
                labels.append(path_name)

    return images, labels


# 從指定路徑讀取訓練數據
def load_dataset(path_name):
    images, labels = read_path(path_name)

    # 將輸入的所有圖片轉成四維陣列，尺寸為(圖片數量 * IMAGE_SIZE * IMAGE_SIZE * 3)
    # 假設兩種類別一共1200張圖，IMAGE_SIZE=64，尺寸 --> 1200 * 64 * 64 * 3
    # 圖片是64 * 64像素, 一個像素3三個顏色值(RGB)
    images = np.array(images)
    print(images.shape)

    # 標註數據，'me'資料夾下都是類別1的圖片，全部指定為0，另外一個資料夾下是類別2，全部指定為1
    labels = np.array([0 if label.endswith('me') else 1 for label in labels])

    return images, labels


def imageChange(cn):
    img = cv2.imread(path + "kg_out.png")
    waveImage = resize_image(img)
    waveImage2 = cv2.resize(waveImage, (224, 224))

    cv2.imshow('img', waveImage2)	#顯示圖片
    cv2.imwrite('D:/chung_yuan/vibration/fall_0412/output.png', waveImage2)

    cv2.waitKey(0)
    cv2.destroyAllWindows()

    load_dataset(path)

def readexcel():
    workbook = openpyxl.load_workbook('fall_data.xlsx')
    sheet = workbook['Sheet']

    rows = sheet.rows
    test_label = []
    test_data = []

    for row in list(rows):
        test_label.append(row[0].value)
        temp_data = []
        for k in range(1, len(row)):
            temp_data.append(row[k].value)
        test_data.append(temp_data)
		
    # 2)數據集劃分
    X_train, X_test, y_train, y_test = train_test_split(test_data, test_label, test_size=0.3, random_state=1)  
	
    # 3)特徵工程 --- 標準化
    transfer = StandardScaler()
    X_train = transfer.fit_transform(X_train)
    X_test = transfer.transform(X_test)

    return transfer, X_train, X_test, y_train, y_test

def remove(sheet, row):
    # iterate the row object
    for cell in row:
		# check the value of each cell in
        # the row, if any of the value is not
        # None return without removing the row
        if cell.value == None:
			return
    # get the row number from the first cell
    # and remove the row
    sheet.delete_rows(row[0].row, 1)


# 與APP建立連線
try:
    s = socket.socket(socket.AF_INET, socket.SOCK_STREAM)  	# 創建socket
    s.bind((HOST, PORT)) 	 # 綁定
    s.listen(1)  	# 監聽
    conn, addr = s.accept()
    # 進入無窮迴圈等代客戶端連線

    while True:
        while ser.in_waiting:  				# 若收到序列資料…
            data_raw = ser.readline()  		# 讀取一行
            data = data_raw.decode()  		# 用預設的UTF-8解碼
            print('接收到的原始資料：', data_raw)
            print('接收到的資料：', data)
            t1 = time.time()
            print(type(data))

            if (int(data) != 0):
                while True:
                    # enter your file path
                    path = './test.xlsx'

                    # load excel file
                    book = openpyxl.load_workbook(path)

                    # select the sheet
                    sheet = book['Sheet']

                    for row in sheet:
                        remove(sheet, row)

                    path = './test.xlsx'
                    book.save(path)

                    arr.append(int(data))
                    t2 = time.time()
                    print('資料：', data)
                    data = data_raw.decode()  # 用預設的UTF-8解碼
                    t2 = time.time()
                    if t2 - t1 > 6.0:
                        print(len(arr))
                        workbook = openpyxl.load_workbook('fall_data.xlsx')
                        sheet = workbook['Sheet']

                        rows = sheet.rows
                        test_label = []
                        test_data = []

                        for row in list(rows):
                            test_label.append(row[0].value)
                            temp_data = []
                            for k in range(1, len(row)):
                                temp_data.append(row[k].value)
                            test_data.append(temp_data)
							
                        # 2)數據集劃分
                        X_train, X_test, y_train, y_test = train_test_split(test_data, test_label, test_size=0.3,
                                                                            random_state=1) 
																			
                        # 3)特徵工程 --- 標準化
                        transfer = StandardScaler()
                        X_train = transfer.fit_transform(X_train)
                        X_test = transfer.transform(X_test)

                        wb = openpyxl.load_workbook('test.xlsx')
                        ws = wb['Sheet']
                        list1 = []
                        list1.append(-1)	#chair->0; fall->1 
                        for i in range(len(arr)):
                            list1.append(arr[i])
                        for i in range(len(arr),487):	#若不足則後面補0
                            list1.append(0)
                        ws.append(list1)
                        wb.save('test.xlsx')	# 儲存檔案 

                        workbook = openpyxl.load_workbook('test.xlsx')
                        sheet = workbook['Sheet']
                        rows = sheet.rows
                        label = []
                        data = []
                        for row in list(rows):
                            label.append(row[0].value)
                            temp_data = []
                            for k in range(1, len(row)):
                                temp_data.append(row[k].value)
                            data.append(temp_data)
                        data = transfer.transform(data)

                        # 4)KNN estimator流程
                        KNN = KNeighborsClassifier(n_neighbors=3, p=2, weights='distance', algorithm='brute') 
                        KNN.fit(X_train, y_train)

                        if KNN.predict(data)[0] == 1:
                            conn.send("1".encode())
                            print(KNN.predict(data)[0])
							
                        arr.clear()
                        break;
                    data_raw = ser.readline()  # 讀取一行
                    data = data_raw.decode()  # 用預設的UTF-8解碼

except KeyboardInterrupt:
    ser.close()  # 清除序列通訊物件
    print('再見！')